from selenium import webdriver
from selenium.webdriver.support.wait import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.webdriver.support.ui import Select
from selenium.webdriver.common.keys import Keys
from selenium.webdriver.common.by import By
from selenium.webdriver.chrome.options import Options
import openpyxl
import xlsxwriter
import time


driver = webdriver.Firefox()
driver.get("http://localhost/ProyectoErasmus/index.php?menu=registro")
elem = driver.find_elements(By.TAG_NAME, "input")
elemSelec = driver.find_elements(By.TAG_NAME, "select")

elem[0].get_attribute("class")

def comprueba(vectorVal, vectorRes, dataframe1, row, elemSelec):
    elem[0].send_keys(vectorVal[0])
    elem[1].send_keys(vectorVal[1])
    elem[2].send_keys(vectorVal[2])
    elem[3].send_keys(vectorVal[3])
    elem[4].send_keys(vectorVal[4])
    elem[5].send_keys(vectorVal[5])
    elem[6].send_keys(vectorVal[6])
    elem[7].send_keys(vectorVal[8])
    elem[8].send_keys(vectorVal[9])
    elem[9].send_keys(vectorVal[10])

    elem[11].click()
    select = Select (elemSelec[0])
    select.select_by_index(vectorVal[7])

    error=driver.find_elements(By.CLASS_NAME,"invalido")

    for i in range(len(vectorVal)):
        if vectorRes[i]=="Valida":
            if i<len(error):

                columnas = dataframe1.max_column+1
                if(len(vectorVal)+i)<= dataframe1.max_column:
                    dataframe1.cell(row=row +1, column=columnas, value='Error')

            else:
                columnas = dataframe1.max_column+1
                if(len(vectorVal)+i)<= dataframe1.max_column:
                    print("Escribe")
                    dataframe1.cell(row=row +1, column=columnas, value=vectorVal[i])
        else:
            columnas = dataframe1.max_column+1
            if(len(vectorVal)+i)<= dataframe1.max_column:
                dataframe1.cell(row=row +1, column=columnas, value='')



dataframe = openpyxl.load_workbook("prueba.xlsx")

dataframe1 = dataframe.active

for row in range(0, dataframe1.max_row):
    vector = []
    columna = 0
    respuesta = []
    for col in dataframe1.iter_cols(1, dataframe1.max_column):
        if (columna < 11):
            
            vector.append("" +str(col[row].value))
            columna = columna + 1

        else:
            respuesta.append("" + str(col[row].value))
            columna = columna + 1
    driver.execute_script("document.forms[0].reset()")
    comprueba(vector, respuesta,dataframe1,row, elemSelec)
    time.sleep(2)

dataframe.save("prueba.xlsx")
time.sleep(5)